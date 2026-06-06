from flask import Blueprint, render_template, request, flash, send_file, jsonify
from flask_login import login_required, current_user
from models import db, Income, Expense, Saving
from balance import compute_balance
from datetime import datetime, date, timedelta
from decimal import Decimal
from sqlalchemy import func, and_
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/')
@login_required
def index():
    return render_template('admin/reports.html')

@reports_bp.route('/data', methods=['GET'])
@login_required
def get_report_data():
    report_type = request.args.get('type', 'month')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    month_str = request.args.get('month')
    category = request.args.get('category')
    source = request.args.get('source')
    
    # Parse dates
    if report_type == 'month' and month_str:
        try:
            year, month = map(int, month_str.split('-'))
            start_date = date(year, month, 1)
            if month == 12:
                end_date = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(year, month + 1, 1) - timedelta(days=1)
        except:
            start_date = date.today().replace(day=1)
            end_date = date.today()
    elif start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        start_date = date.today().replace(day=1)
        end_date = date.today()
    
    # Income data
    income_query = Income.query.filter(
        Income.user_id == current_user.id,
        Income.is_deleted == False,
        Income.date >= start_date,
        Income.date <= end_date
    )
    if source:
        income_query = income_query.filter(Income.source == source)
    incomes = income_query.all()
    
    # Expense data
    expense_query = Expense.query.filter(
        Expense.user_id == current_user.id,
        Expense.is_deleted == False,
        Expense.date >= start_date,
        Expense.date <= end_date
    )
    if category:
        expense_query = expense_query.filter(Expense.category == category)
    expenses = expense_query.all()
    
    # Savings data
    savings = Saving.query.filter(
        Saving.user_id == current_user.id,
        Saving.is_deleted == False,
        Saving.date >= start_date,
        Saving.date <= end_date
    ).all()
    
    # Summary
    total_income = sum(i.amount for i in incomes) if incomes else Decimal('0')
    total_expenses = sum(e.amount for e in expenses) if expenses else Decimal('0')
    
    balance = compute_balance(current_user.id, end_date)
    
    # Income by source
    income_by_source = {}
    for inc in incomes:
        income_by_source[inc.source] = income_by_source.get(inc.source, Decimal('0')) + inc.amount
    
    # Expenses by category
    expense_by_category = {}
    for exp in expenses:
        expense_by_category[exp.category] = expense_by_category.get(exp.category, Decimal('0')) + exp.amount
    
    # Savings by category
    savings_by_category = {}
    for sav in savings:
        if not sav.is_recovered or sav.category != 'Borrowed to Others':
            savings_by_category[sav.category] = savings_by_category.get(sav.category, Decimal('0')) + sav.amount
    
    report_data = {
        'start_date': start_date.isoformat(),
        'end_date': end_date.isoformat(),
        'total_income': float(total_income),
        'total_expenses': float(total_expenses),
        'net_surplus': float(total_income - total_expenses),
        'net_available': float(balance['net_available']),
        'total_savings': float(balance['total_savings']),
        'income_by_source': {k: float(v) for k, v in income_by_source.items()},
        'expense_by_category': {k: float(v) for k, v in expense_by_category.items()},
        'savings_by_category': {k: float(v) for k, v in savings_by_category.items()},
        'incomes': [{'date': i.date.isoformat(), 'source': i.source, 'amount': float(i.amount), 'remarks': i.remarks} for i in incomes],
        'expenses': [{'date': e.date.isoformat(), 'category': e.category, 'description': e.description, 'amount': float(e.amount)} for e in expenses],
        'savings': [{'date': s.date.isoformat(), 'category': s.category, 'sub_category': s.sub_category, 'amount': float(s.amount), 'remarks': s.remarks, 'recovered': s.is_recovered} for s in savings]
    }
    
    return jsonify(report_data)

@reports_bp.route('/download/pdf')
@login_required
def download_pdf():
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import io
    import base64
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER
    
    # Get report data
    report_type = request.args.get('type', 'month')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    month_str = request.args.get('month')
    
    # Parse dates (same as above)
    if report_type == 'month' and month_str:
        try:
            year, month = map(int, month_str.split('-'))
            start_date = date(year, month, 1)
            if month == 12:
                end_date = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(year, month + 1, 1) - timedelta(days=1)
        except:
            start_date = date.today().replace(day=1)
            end_date = date.today()
    elif start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        start_date = date.today().replace(day=1)
        end_date = date.today()
    
    # Generate charts as images
    # Income pie chart
    incomes = Income.query.filter(
        Income.user_id == current_user.id,
        Income.is_deleted == False,
        Income.date >= start_date,
        Income.date <= end_date
    ).all()
    
    income_by_source = {}
    for inc in incomes:
        income_by_source[inc.source] = income_by_source.get(inc.source, 0) + float(inc.amount)
    
    fig1, ax1 = plt.subplots(figsize=(6, 4))
    if income_by_source:
        ax1.pie(income_by_source.values(), labels=income_by_source.keys(), autopct='%1.1f%%', startangle=90)
        ax1.set_title('Income Distribution by Source')
    else:
        ax1.text(0.5, 0.5, 'No income data', ha='center', va='center')
        ax1.set_title('Income Distribution')
    
    img1 = io.BytesIO()
    plt.savefig(img1, format='png', bbox_inches='tight', dpi=100)
    img1.seek(0)
    income_chart = Image(img1, width=5*inch, height=3.5*inch)
    plt.close()
    
    # Expense pie chart
    expenses = Expense.query.filter(
        Expense.user_id == current_user.id,
        Expense.is_deleted == False,
        Expense.date >= start_date,
        Expense.date <= end_date
    ).all()
    
    expense_by_category = {}
    for exp in expenses:
        expense_by_category[exp.category] = expense_by_category.get(exp.category, 0) + float(exp.amount)
    
    fig2, ax2 = plt.subplots(figsize=(6, 4))
    if expense_by_category:
        ax2.pie(expense_by_category.values(), labels=expense_by_category.keys(), autopct='%1.1f%%', startangle=90)
        ax2.set_title('Expense Distribution by Category')
    else:
        ax2.text(0.5, 0.5, 'No expense data', ha='center', va='center')
        ax2.set_title('Expense Distribution')
    
    img2 = io.BytesIO()
    plt.savefig(img2, format='png', bbox_inches='tight', dpi=100)
    img2.seek(0)
    expense_chart = Image(img2, width=5*inch, height=3.5*inch)
    plt.close()
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, textColor=colors.HexColor('#2c3e50'), alignment=TA_CENTER, spaceAfter=30)
    story.append(Paragraph("FinTrack Financial Report", title_style))
    
    # Date range
    date_style = ParagraphStyle('DateStyle', parent=styles['Normal'], fontSize=12, textColor=colors.HexColor('#7f8c8d'), alignment=TA_CENTER, spaceAfter=20)
    story.append(Paragraph(f"Period: {start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}", date_style))
    
    # Summary table
    balance = compute_balance(current_user.id, end_date)
    summary_data = [
        ['Metric', 'Amount (₹)'],
        ['Total Income', f"{float(balance['total_income']):,.2f}"],
        ['Total Expenses', f"{float(balance['total_expenses']):,.2f}"],
        ['Net Surplus', f"{float(balance['total_income'] - balance['total_expenses']):,.2f}"],
        ['Borrowed Out (Unrecovered)', f"{float(balance['total_borrowed_out_unrecovered']):,.2f}"],
        ['Net Available Balance', f"{float(balance['net_available']):,.2f}"],
        ['Total Savings/Assets', f"{float(balance['total_savings']):,.2f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Charts
    story.append(Paragraph("Income Distribution", styles['Heading2']))
    story.append(income_chart)
    story.append(Spacer(1, 20))
    
    story.append(Paragraph("Expense Distribution", styles['Heading2']))
    story.append(expense_chart)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    return send_file(buffer, as_attachment=True, download_name=f"fintrack_report_{start_date}_{end_date}.pdf", mimetype='application/pdf')

@reports_bp.route('/download/excel')
@login_required
def download_excel():
    # Get report data (same as above)
    report_type = request.args.get('type', 'month')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    month_str = request.args.get('month')
    
    if report_type == 'month' and month_str:
        try:
            year, month = map(int, month_str.split('-'))
            start_date = date(year, month, 1)
            if month == 12:
                end_date = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(year, month + 1, 1) - timedelta(days=1)
        except:
            start_date = date.today().replace(day=1)
            end_date = date.today()
    elif start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    else:
        start_date = date.today().replace(day=1)
        end_date = date.today()
    
    wb = Workbook()
    
    # Income sheet
    ws1 = wb.active
    ws1.title = "Income"
    headers = ['Date', 'Source', 'Amount (₹)', 'Remarks']
    ws1.append(headers)
    
    incomes = Income.query.filter(
        Income.user_id == current_user.id,
        Income.is_deleted == False,
        Income.date >= start_date,
        Income.date <= end_date
    ).order_by(Income.date.desc()).all()
    
    for inc in incomes:
        ws1.append([inc.date, inc.source, float(inc.amount), inc.remarks])
    
    # Expenses sheet
    ws2 = wb.create_sheet("Expenses")
    ws2.append(['Date', 'Category', 'Description', 'Amount (₹)'])
    
    expenses = Expense.query.filter(
        Expense.user_id == current_user.id,
        Expense.is_deleted == False,
        Expense.date >= start_date,
        Expense.date <= end_date
    ).order_by(Expense.date.desc()).all()
    
    for exp in expenses:
        ws2.append([exp.date, exp.category, exp.description, float(exp.amount)])
    
    # Savings sheet
    ws3 = wb.create_sheet("Savings")
    ws3.append(['Date', 'Category', 'Sub-Category', 'Amount (₹)', 'Remarks', 'Recovered'])
    
    savings = Saving.query.filter(
        Saving.user_id == current_user.id,
        Saving.is_deleted == False,
        Saving.date >= start_date,
        Saving.date <= end_date
    ).order_by(Saving.date.desc()).all()
    
    for sav in savings:
        ws3.append([sav.date, sav.category, sav.sub_category or '-', float(sav.amount), sav.remarks, 'Yes' if sav.is_recovered else 'No'])
    
    # Summary sheet
    ws4 = wb.create_sheet("Summary")
    balance = compute_balance(current_user.id, end_date)
    summary_data = [
        ['Financial Summary', ''],
        ['Period', f"{start_date} to {end_date}"],
        ['', ''],
        ['Metric', 'Amount (₹)'],
        ['Total Income', float(balance['total_income'])],
        ['Total Expenses', float(balance['total_expenses'])],
        ['Net Surplus', float(balance['total_income'] - balance['total_expenses'])],
        ['Borrowed Out (Unrecovered)', float(balance['total_borrowed_out_unrecovered'])],
        ['Net Available Balance', float(balance['net_available'])],
        ['Total Savings/Assets', float(balance['total_savings'])],
    ]
    
    for row in summary_data:
        ws4.append(row)
    
    # Style sheets
    for sheet in [ws1, ws2, ws3]:
        # Format headers
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-fit columns
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Format currency columns
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                if 'Amount' in str(sheet.cell(1, cell.column).value):
                    cell.number_format = '#,##0.00'
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(buffer, as_attachment=True, download_name=f"fintrack_report_{start_date}_{end_date}.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')